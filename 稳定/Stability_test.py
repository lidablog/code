#coding:utf-8
import openpyxl
import pandas as pd
import os
import matplotlib as plt
import xlsxwriter
from PIL import ImageGrab
import win32com.client as win32
huan=input("请输入要统计的deb版本数量:")
huan=int(huan)
dds=0
while (dds<huan):
  #matplotlib.use('Agg')
  df=pd.read_excel("Cinderella稳定性测试结果追踪.xlsx",sheet_name="Cinderella稳定性测试结果追踪",usecols=[1,3,13,14,17,20,21])
  df.to_excel("dd.xlsx")
  version=input("请输入要统计的deb版本号:")
  se=pd.read_excel("dd.xlsx")
  te=se[se["deb版本"].str.contains(version,na=False)]
  #挑出HF1.1的
  te.to_excel("ee.xlsx")
  #HF1.1的文件
  ff=te["日期"].value_counts()
  #日期重复个数
  ff.to_excel("ff.xlsx")
  gg=pd.read_excel("ff.xlsx")
  hh=gg.sort_values(by="Unnamed: 0", ascending=True)
  #按时间排序
  ii=hh.reset_index(drop=True)
  #索引重排序
  ii.to_excel("hh.xlsx")
  #输出时间-总任务数文件
  kk=pd.read_excel("ee.xlsx",usecols=[2,4])
  kk.to_excel("kk.xlsx")
  time=kk.groupby("日期").sum()
  time.to_excel("time.xlsx")
  nn=pd.read_excel("time.xlsx")
  nn.to_excel("time.xlsx")
  #输出时间-总清扫时间
  ll=pd.read_excel("ee.xlsx",usecols=[2,7])
  ll.to_excel("ll.xlsx")
  block=ll.groupby("日期").sum()
  block.to_excel("block.xlsx")
  oo=pd.read_excel("block.xlsx")
  oo.to_excel("block.xlsx")
  #总清扫中断异常
  mm=pd.read_excel("ee.xlsx",usecols=[2,8])
  mm.to_excel("mm.xlsx")
  a_block=mm.groupby("日期").sum()
  a_block.to_excel("a_block.xlsx")
  pp=pd.read_excel("a_block.xlsx")
  pp.to_excel("a_block.xlsx")
  #总软件清扫中断
  #num=pd.read_excel("处理后.xlsx")
  #num.to_excel("num.xlsx")
  #总任务数
  qa=pd.read_excel("ee.xlsx",usecols=[2,7])
  qa.to_excel("qa.xlsx")
  qa["异常中断"][qa.异常中断>0]=1
  qa.to_excel("111.xlsx")
  q_block=qa.groupby("日期").sum()
  q_block.to_excel("q_block.xlsx")
  xx=pd.read_excel("q_block.xlsx")
  xx.to_excel("q_block.xlsx")
  #总清扫中断异常权重
  ws=pd.read_excel("ee.xlsx",usecols=[2,8])
  ws.to_excel("ws.xlsx")
  ws["软件中断"][ws.软件中断>0]=1
  q_a_block=ws.groupby("日期").sum()
  q_a_block.to_excel("q_a_block.xlsx")
  yy=pd.read_excel("q_a_block.xlsx")
  yy.to_excel("q_a_block.xlsx")
  #总清扫中断软件异常权重
  def chek(filename):
      bk = openpyxl.load_workbook(filename)
      sheet = bk.active
      minrow = sheet.min_row
      maxrow = sheet.max_row
      date_col = [sheet.cell(n,2).value for n in range(minrow,maxrow+1)][1:]
      total_col = [sheet.cell(n,3).value for n in range(minrow,maxrow+1)][1:]
      new_date = date_col[::5]
      new_total = total_col[::5]
      nds = []
      for nd in date_col:
          index = date_col.index(nd)
          tmp = date_col[index:index+5]
          tmp_t = total_col[index:index+5]
          tmp_tt = sum(tmp_t)
          index_ = str(tmp[0])[-4:]+"-"+str(tmp[-1])[-4:]
          if len(tmp_t)==5:
              nds.append([index_,tmp_tt])
      return nds

  def write_xls(data):
      bk = openpyxl.Workbook()
      sheet = bk.active
      sheet.cell(1,1).value="时间段"
      sheet.cell(1,2).value="总任务数"
      for i in range(len(data)):
          sheet.append(data[i])
      bk.save('处理后.xlsx')

            

  data = chek('hh.xlsx')
  write_xls(data)
  #demo

  def aa(filename):
      bk = openpyxl.load_workbook(filename)
      sheet = bk.active
      minrow = sheet.min_row
      maxrow = sheet.max_row
      date_col = [sheet.cell(n,2).value for n in range(minrow,maxrow+1)][1:]
      total_col = [sheet.cell(n,3).value for n in range(minrow,maxrow+1)][1:]
      new_date = date_col[::5]
      new_total = total_col[::5]
      nds = []
      for nd in date_col:
          index = date_col.index(nd)
          tmp = date_col[index:index+5]
          tmp_t = total_col[index:index+5]
          tmp_tt = sum(tmp_t)
          index_ = str(tmp[0])[-4:]+"-"+str(tmp[-1])[-4:]
          if len(tmp_t)==5:
              nds.append([index_,tmp_tt])
      return nds

  def bb(data):
      bk = openpyxl.Workbook()
      sheet = bk.active
      sheet.cell(1,1).value="时间段"
      sheet.cell(1,2).value="总时间数"
      for i in range(len(data)):
          sheet.append(data[i])
      bk.save('time_rm.xlsx')

            

  data = aa('time.xlsx')
  bb(data)
  #demo2
  def cc(filename):
      bk = openpyxl.load_workbook(filename)
      sheet = bk.active
      minrow = sheet.min_row
      maxrow = sheet.max_row
      date_col = [sheet.cell(n,2).value for n in range(minrow,maxrow+1)][1:]
      total_col = [sheet.cell(n,3).value for n in range(minrow,maxrow+1)][1:]
      new_date = date_col[::5]
      new_total = total_col[::5]
      nds = []
      for nd in date_col:
          index = date_col.index(nd)
          tmp = date_col[index:index+5]
          tmp_t = total_col[index:index+5]
          tmp_tt = sum(tmp_t)
          index_ = str(tmp[0])[-4:]+"-"+str(tmp[-1])[-4:]
          if len(tmp_t)==5:
              nds.append([index_,tmp_tt])
      return nds

  def dd(data):
      bk = openpyxl.Workbook()
      sheet = bk.active
      sheet.cell(1,1).value="时间段"
      sheet.cell(1,2).value="异常中断总数"
      for i in range(len(data)):
          sheet.append(data[i])
      bk.save('block_rm.xlsx')

            

  data = cc('block.xlsx')
  dd(data)
  #demo3
  def ee(filename):
      bk = openpyxl.load_workbook(filename)
      sheet = bk.active
      minrow = sheet.min_row
      maxrow = sheet.max_row
      date_col = [sheet.cell(n,2).value for n in range(minrow,maxrow+1)][1:]
      total_col = [sheet.cell(n,3).value for n in range(minrow,maxrow+1)][1:]
      new_date = date_col[::5]
      new_total = total_col[::5]
      nds = []
      for nd in date_col:
          index = date_col.index(nd)
          tmp = date_col[index:index+5]
          tmp_t = total_col[index:index+5]
          tmp_tt = sum(tmp_t)
          index_ = str(tmp[0])[-4:]+"-"+str(tmp[-1])[-4:]
          if len(tmp_t)==5:
              nds.append([index_,tmp_tt])
      return nds

  def ff(data):
      bk = openpyxl.Workbook()
      sheet = bk.active
      sheet.cell(1,1).value="时间段"
      sheet.cell(1,2).value="软件中断总数"
      for i in range(len(data)):
          sheet.append(data[i])
      bk.save('a_block_rm.xlsx')

            

  data = ee('a_block.xlsx')
  ff(data)
  #demo4
  def gg(filename):
      bk = openpyxl.load_workbook(filename)
      sheet = bk.active
      minrow = sheet.min_row
      maxrow = sheet.max_row
      date_col = [sheet.cell(n,2).value for n in range(minrow,maxrow+1)][1:]
      total_col = [sheet.cell(n,3).value for n in range(minrow,maxrow+1)][1:]
      new_date = date_col[::5]
      new_total = total_col[::5]
      nds = []
      for nd in date_col:
          index = date_col.index(nd)
          tmp = date_col[index:index+5]
          tmp_t = total_col[index:index+5]
          tmp_tt = sum(tmp_t)
          index_ = str(tmp[0])[-4:]+"-"+str(tmp[-1])[-4:]
          if len(tmp_t)==5:
              nds.append([index_,tmp_tt])
      return nds

  def hh(data):
      bk = openpyxl.Workbook()
      sheet = bk.active
      sheet.cell(1,1).value="时间段"
      sheet.cell(1,2).value="异常中断总数权重"
      for i in range(len(data)):
          sheet.append(data[i])
      bk.save('q_block_rm.xlsx')

            

  data = gg('q_block.xlsx')
  hh(data)
  #demo5
  def ii(filename):
      bk = openpyxl.load_workbook(filename)
      sheet = bk.active
      minrow = sheet.min_row
      maxrow = sheet.max_row
      date_col = [sheet.cell(n,2).value for n in range(minrow,maxrow+1)][1:]
      total_col = [sheet.cell(n,3).value for n in range(minrow,maxrow+1)][1:]
      new_date = date_col[::5]
      new_total = total_col[::5]
      nds = []
      for nd in date_col:
          index = date_col.index(nd)
          tmp = date_col[index:index+5]
          tmp_t = total_col[index:index+5]
          tmp_tt = sum(tmp_t)
          index_ = str(tmp[0])[-4:]+"-"+str(tmp[-1])[-4:]
          if len(tmp_t)==5:
              nds.append([index_,tmp_tt])
      return nds

  def jj(data):
      bk = openpyxl.Workbook()
      sheet = bk.active
      sheet.cell(1,1).value="时间段"
      sheet.cell(1,2).value="软件中断总数权重"
      for i in range(len(data)):
          sheet.append(data[i])
      bk.save('q_a_block_rm.xlsx')

            

  data = ii('q_a_block.xlsx')
  jj(data)
  #demo6


  #这里加demo
  num=pd.read_excel("处理后.xlsx")
  num.to_excel("num.xlsx")
  #总任务数
  ss=pd.read_excel("time_rm.xlsx")
  tt=pd.read_excel("block_rm.xlsx")
  uu=pd.read_excel("a_block_rm.xlsx")
  ed=pd.read_excel("q_block_rm.xlsx")
  rf=pd.read_excel("q_a_block_rm.xlsx")
  qq=pd.merge(ss,tt)
  rr=pd.merge(qq,uu)
  tt=pd.merge(rr,num)
  tg=pd.merge(tt,ed)
  yh=pd.merge(tg,rf)
  yh.to_excel("all.xlsx")
  #合并
  vv=pd.read_excel("all.xlsx")
  #每小时异常数
  vv["每小时异常数"]=round(vv["异常中断总数"]/(vv["总时间数"]/60),2)
  vv["每小时(软件异常数)"]=round(vv["软件中断总数"]/(vv["总时间数"]/60),2)
  #成功率
  vv["成功率"]=round(((vv["总任务数"]-vv["异常中断总数权重"])/(vv["总任务数"]))*100,2)
  vv["成功率(软件)"]=round(((vv["总任务数"]-vv["软件中断总数权重"])/(vv["总任务数"])*100),2)
  vv.to_excel("东升稳定性测试数据.xlsx")
  #总数据表
  os.remove("111.xlsx")
  os.remove("a_block.xlsx")
  os.remove("a_block_rm.xlsx")
  os.remove("all.xlsx")
  os.remove("block.xlsx")
  os.remove("block_rm.xlsx")
  os.remove("dd.xlsx")
  os.remove("ee.xlsx")
  os.remove("ff.xlsx")
  os.remove("hh.xlsx")
  os.remove("kk.xlsx")
  os.remove("ll.xlsx")
  os.remove("mm.xlsx")
  os.remove("num.xlsx")
  os.remove("q_a_block.xlsx")
  os.remove("q_a_block_rm.xlsx")
  os.remove("q_block.xlsx")
  os.remove("q_block_rm.xlsx")
  os.remove("qa.xlsx")
  os.remove("time.xlsx")
  os.remove("time_rm.xlsx")
  os.remove("ws.xlsx")
  os.remove("处理后.xlsx")



  if version=="HF1.1":
     name_1=version + "成功率"
     name_2=version + "每小时异常数"
  elif version=="HF2.0":
     name_1=version + "成功率"
     name_2=version + "每小时异常数"
  else:
     print("无此版本")

  #vv.to_excel("东升稳定性测试数据1.xlsx")

  #a = pd.read_csv('东升稳定性测试数据1.xlsx',columns_col=0)
  a=len(vv)
  writer = pd.ExcelWriter('东升稳定性测试数据.xlsx', engine='xlsxwriter')
  vv.to_excel(writer, sheet_name='Sheet1')
  workbook  = writer.book
  worksheet = writer.sheets['Sheet1']
  chart = workbook.add_chart({'type': 'line'})
  chart.set_title({'name': name_1})
  chart.add_series({
      'categories': ['sheet1',1,2,a-1,2], # x轴显示内容
      'values':     ['sheet1',1,11,a-1,11],
      'line':       {'color': 'red'}, # 线条颜色
      'name':       '成功率(%)', # 图例名称
  })
  chart.add_series({
      'categories': ['sheet1',1,2,a-1,2], # x轴显示内容
      'values':     ['sheet1',1,12,a-1,12],
      'line':       {'color': 'blue'}, # 线条颜色
      'name':       '软件成功率(%)', # 图例名称
  })
  chart.set_y_axis({'min': 0, 'max': 100})
  chart.set_size({'width': 2300, 'height': 570})
  chart.set_table()
  worksheet.insert_chart('O2', chart)
  #writer.save()
  chart = workbook.add_chart({'type': 'line'})
  chart.set_title({'name':name_2})
  chart.add_series({
      'categories': ['sheet1',1,2,a-1,2], # x轴显示内容
      'values':     ['sheet1',1,9,a-1,9],
      'line':       {'color': 'red'}, # 线条颜色
      'name':       '每小时异常数', # 图例名称
  })
  chart.add_series({
      'categories': ['sheet1',1,2,a-1,2], # x轴显示内容
      'values':     ['sheet1',1,10,a-1,10],
      'line':       {'color': 'blue'}, # 线条颜色
      'name':       '每小时异常数(软件)', # 图例名称
  })
  chart.set_y_axis({'min': 0, 'max': 0.5})
  chart.set_size({'width': 2300, 'height': 570})
  chart.set_table()
  worksheet.insert_chart('O2', chart)
  writer.save()
  #制图

  path=os.getcwd()
  excel = win32.gencache.EnsureDispatch('Excel.Application')
  workbook = excel.Workbooks.Open(path+'\东升稳定性测试数据.xlsx')
  version_1=version
  num="I"
  for sheet in workbook.Worksheets:
      for i, shape in enumerate(sheet.Shapes):
          if shape.Name.startswith('Chart'):
              shape.Copy()
              image = ImageGrab.grabclipboard()   
              image.convert('RGB').save(path+'\{}.jpg'.format(version+"_"+num), 'jpeg')
              num=num+"I"
  print("输出成功"+version_1+"数据")
  excel.Quit()
  dds=dds+1
  #保存出来

  #os.remove("Cinderella稳定性测试结果追踪.xlsx")










































